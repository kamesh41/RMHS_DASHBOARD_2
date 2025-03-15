from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from .models import Area, Shift, Material, FeedingOperation, ReclaimingOperation, CrushingOperation, ReceivingOperation, DailySummary
from datetime import datetime, timedelta

def dashboard(request):
    """Main operations dashboard view"""
    # Get filter parameters
    date_str = request.GET.get('date')
    area_id = request.GET.get('area')
    shift_name = request.GET.get('shift')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get all areas for filter dropdown
    areas = Area.objects.all()
    selected_area = None
    if area_id:
        selected_area = get_object_or_404(Area, id=area_id)
    
    # Get shifts for the selected date
    shifts = Shift.objects.filter(date=selected_date).order_by('name')
    if not shifts.exists():
        # Create shifts for this date if they don't exist
        for shift_name in ['A', 'B', 'C']:
            Shift.objects.create(name=shift_name, date=selected_date)
        shifts = Shift.objects.filter(date=selected_date).order_by('name')
    
    selected_shift = shift_name
    
    # Build base querysets
    feeding_qs = FeedingOperation.objects.filter(shift__date=selected_date)
    reclaiming_qs = ReclaimingOperation.objects.filter(shift__date=selected_date)
    crushing_qs = CrushingOperation.objects.filter(shift__date=selected_date)
    receiving_qs = ReceivingOperation.objects.filter(shift__date=selected_date)
    
    # Apply area filter if selected
    if selected_area:
        feeding_qs = feeding_qs.filter(area=selected_area)
        reclaiming_qs = reclaiming_qs.filter(area=selected_area)
        crushing_qs = crushing_qs.filter(area=selected_area)
        receiving_qs = receiving_qs.filter(area=selected_area)
    
    # Apply shift filter if selected
    if selected_shift:
        feeding_qs = feeding_qs.filter(shift__name=selected_shift)
        reclaiming_qs = reclaiming_qs.filter(shift__name=selected_shift)
        crushing_qs = crushing_qs.filter(shift__name=selected_shift)
        receiving_qs = receiving_qs.filter(shift__name=selected_shift)
    
    # Calculate totals
    feeding_total = feeding_qs.aggregate(total=Sum('quantity'))['total'] or 0
    reclaiming_total = reclaiming_qs.aggregate(total=Sum('quantity'))['total'] or 0
    crushing_total = crushing_qs.aggregate(total=Sum('quantity'))['total'] or 0
    receiving_total = receiving_qs.aggregate(total=Sum('quantity'))['total'] or 0
    
    # Get all unique materials used in any operation for this date/filter
    all_materials = Material.objects.filter(
        Q(feedingoperation__shift__date=selected_date) |
        Q(reclaimingoperation__shift__date=selected_date) |
        Q(crushingoperation__shift__date=selected_date) |
        Q(receivingoperation__shift__date=selected_date)
    ).distinct()
    
    if selected_area:
        all_materials = all_materials.filter(
            Q(feedingoperation__area=selected_area) |
            Q(reclaimingoperation__area=selected_area) |
            Q(crushingoperation__area=selected_area) |
            Q(receivingoperation__area=selected_area)
        ).distinct()
    
    if selected_shift:
        all_materials = all_materials.filter(
            Q(feedingoperation__shift__name=selected_shift) |
            Q(reclaimingoperation__shift__name=selected_shift) |
            Q(crushingoperation__shift__name=selected_shift) |
            Q(receivingoperation__shift__name=selected_shift)
        ).distinct()
    
    # Get summary data organized by material type
    materials_by_type = {}
    for material in all_materials:
        type_name = material.type.name if material.type else "Uncategorized"
        if type_name not in materials_by_type:
            materials_by_type[type_name] = []
        
        # Calculate total for this material across all operations
        material_total = 0
        for item in feeding_qs.filter(material=material):
            material_total += item.quantity
        for item in reclaiming_qs.filter(material=material):
            material_total += item.quantity
        for item in crushing_qs.filter(material=material):
            material_total += item.quantity
        for item in receiving_qs.filter(material=material):
            material_total += item.quantity
        
        # Add material with its total to the list
        material_with_total = material
        material_with_total.total = material_total
        materials_by_type[type_name].append(material_with_total)
    
    # Get summary data
    feeding_summary = feeding_qs.values('material__name', 'material__type__name').annotate(total=Sum('quantity')).order_by('material__type__name', 'material__name')
    reclaiming_summary = reclaiming_qs.values('material__name', 'material__type__name').annotate(total=Sum('quantity')).order_by('material__type__name', 'material__name')
    crushing_summary = crushing_qs.values('material__name', 'material__type__name').annotate(total=Sum('quantity')).order_by('material__type__name', 'material__name')
    receiving_summary = receiving_qs.values('material__name', 'material__type__name').annotate(total=Sum('quantity')).order_by('material__type__name', 'material__name')
    
    # Get recent activities (limited to 5)
    recent_feeding = feeding_qs.order_by('-timestamp')[:5]
    recent_reclaiming = reclaiming_qs.order_by('-timestamp')[:5]
    recent_crushing = crushing_qs.order_by('-timestamp')[:5]
    recent_receiving = receiving_qs.order_by('-timestamp')[:5]
    
    context = {
        'selected_date': selected_date,
        'areas': areas,
        'selected_area': selected_area,
        'shifts': shifts,
        'selected_shift': selected_shift,
        'feeding_data': feeding_qs,
        'reclaiming_data': reclaiming_qs,
        'crushing_data': crushing_qs,
        'receiving_data': receiving_qs,
        'feeding_summary': feeding_summary,
        'reclaiming_summary': reclaiming_summary,
        'crushing_summary': crushing_summary,
        'receiving_summary': receiving_summary,
        'feeding_total': feeding_total,
        'reclaiming_total': reclaiming_total,
        'crushing_total': crushing_total,
        'receiving_total': receiving_total,
        'recent_feeding': recent_feeding,
        'recent_reclaiming': recent_reclaiming,
        'recent_crushing': recent_crushing,
        'recent_receiving': recent_receiving,
        'all_materials': all_materials,
        'materials_by_type': materials_by_type,
    }
    
    return render(request, 'operations/dashboard.html', context)

def feeding_view(request):
    """View for feeding data entry and display"""
    # Get filter parameters
    date_str = request.GET.get('date')
    area_id = request.GET.get('area')
    shift_name = request.GET.get('shift')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get all areas and materials for dropdowns
    areas = Area.objects.all()
    materials = Material.objects.all()
    
    # Get shifts for the selected date
    shifts = Shift.objects.filter(date=selected_date).order_by('name')
    if not shifts.exists():
        # Create shifts for this date if they don't exist
        for shift_name in ['A', 'B', 'C']:
            Shift.objects.create(name=shift_name, date=selected_date)
        shifts = Shift.objects.filter(date=selected_date).order_by('name')
    
    selected_area = None
    if area_id:
        selected_area = get_object_or_404(Area, id=area_id)
    
    selected_shift = shift_name
    
    # Handle form submission
    if request.method == 'POST':
        area_id = request.POST.get('area')
        shift_id = request.POST.get('shift')
        destination = request.POST.get('destination')
        notes = request.POST.get('notes')
        reported_by = request.POST.get('reported_by')
        
        # Get material names, IDs, and quantities as lists
        material_names = request.POST.getlist('material_names[]')
        material_ids = request.POST.getlist('material_ids[]')
        quantities = request.POST.getlist('quantities[]')
        
        # Validate form data
        if area_id and shift_id and destination and reported_by and material_names and quantities:
            try:
                area = Area.objects.get(id=area_id)
                shift = Shift.objects.get(id=shift_id)
                
                # Create entries for each material
                success_count = 0
                for i in range(len(material_names)):
                    if material_names[i] and quantities[i]:
                        try:
                            # Try to get material by ID first
                            material = None
                            if material_ids[i]:
                                try:
                                    material = Material.objects.get(id=material_ids[i])
                                except Material.DoesNotExist:
                                    pass
                            
                            # If no material found by ID, try to find by name
                            if not material:
                                material_name = material_names[i].strip()
                                try:
                                    material = Material.objects.get(name__iexact=material_name)
                                except Material.DoesNotExist:
                                    # Create new material if it doesn't exist
                                    material = Material.objects.create(
                                        name=material_name,
                                        is_active=True
                                    )
                                    messages.info(request, f'Created new material: {material_name}')
                            
                            quantity = float(quantities[i])
                            
                            # Create new feeding entry
                            FeedingOperation.objects.create(
                                area=area,
                                shift=shift,
                                material=material,
                                quantity=quantity,
                                destination=destination,
                                notes=notes,
                                reported_by=reported_by
                            )
                            success_count += 1
                        except Exception as e:
                            messages.error(request, f'Error adding material {i+1}: {str(e)}')
                
                if success_count > 0:
                    messages.success(request, f'{success_count} feeding entries added successfully!')
                    return redirect('feeding')
            except Exception as e:
                messages.error(request, f'Error adding feeding entries: {str(e)}')
        else:
            messages.error(request, 'Please fill all required fields')
    
    # Get feeding data for display
    feeding_data = FeedingOperation.objects.filter(shift__date=selected_date)
    
    # Apply filters if provided
    if area_id:
        feeding_data = feeding_data.filter(area_id=area_id)
    
    if shift_name:
        feeding_data = feeding_data.filter(shift__name=shift_name)
    
    # Order by timestamp
    feeding_data = feeding_data.order_by('-timestamp')
    
    context = {
        'selected_date': selected_date,
        'areas': areas,
        'materials': materials,
        'shifts': shifts,
        'selected_area': selected_area,
        'selected_shift': selected_shift,
        'feeding_data': feeding_data,
    }
    
    return render(request, 'operations/feeding.html', context)

def reclaiming_view(request):
    """View for reclaiming data entry and display"""
    # Get filter parameters
    date_str = request.GET.get('date')
    area_id = request.GET.get('area')
    shift_name = request.GET.get('shift')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get all areas and materials for dropdowns
    areas = Area.objects.all()
    materials = Material.objects.all()
    
    # Get shifts for the selected date
    shifts = Shift.objects.filter(date=selected_date).order_by('name')
    if not shifts.exists():
        # Create shifts for this date if they don't exist
        for shift_name in ['A', 'B', 'C']:
            Shift.objects.create(name=shift_name, date=selected_date)
        shifts = Shift.objects.filter(date=selected_date).order_by('name')
    
    selected_area = None
    if area_id:
        selected_area = get_object_or_404(Area, id=area_id)
    
    selected_shift = shift_name
    
    # Handle form submission
    if request.method == 'POST':
        area_id = request.POST.get('area')
        shift_id = request.POST.get('shift')
        source = request.POST.get('source')
        notes = request.POST.get('notes')
        reported_by = request.POST.get('reported_by')
        
        # Get material names, IDs, and quantities as lists
        material_names = request.POST.getlist('material_names[]')
        material_ids = request.POST.getlist('material_ids[]')
        quantities = request.POST.getlist('quantities[]')
        
        # Validate form data
        if area_id and shift_id and source and reported_by and material_names and quantities:
            try:
                area = Area.objects.get(id=area_id)
                shift = Shift.objects.get(id=shift_id)
                
                # Create entries for each material
                success_count = 0
                for i in range(len(material_names)):
                    if material_names[i] and quantities[i]:
                        try:
                            # Try to get material by ID first
                            material = None
                            if material_ids[i]:
                                try:
                                    material = Material.objects.get(id=material_ids[i])
                                except Material.DoesNotExist:
                                    pass
                            
                            # If no material found by ID, try to find by name
                            if not material:
                                material_name = material_names[i].strip()
                                try:
                                    material = Material.objects.get(name__iexact=material_name)
                                except Material.DoesNotExist:
                                    # Create new material if it doesn't exist
                                    material = Material.objects.create(
                                        name=material_name,
                                        is_active=True
                                    )
                                    messages.info(request, f'Created new material: {material_name}')
                            
                            quantity = float(quantities[i])
                            
                            # Create new reclaiming entry
                            ReclaimingOperation.objects.create(
                                area=area,
                                shift=shift,
                                material=material,
                                quantity=quantity,
                                source=source,
                                notes=notes,
                                reported_by=reported_by
                            )
                            success_count += 1
                        except Exception as e:
                            messages.error(request, f'Error adding material {i+1}: {str(e)}')
                
                if success_count > 0:
                    messages.success(request, f'{success_count} reclaiming entries added successfully!')
                return redirect('reclaiming')
            except Exception as e:
                messages.error(request, f'Error adding reclaiming entries: {str(e)}')
        else:
            messages.error(request, 'Please fill all required fields')
    
    # Get reclaiming data for display
    reclaiming_data = ReclaimingOperation.objects.filter(shift__date=selected_date)
    
    # Apply filters if provided
    if area_id:
        reclaiming_data = reclaiming_data.filter(area_id=area_id)
    
    if shift_name:
        reclaiming_data = reclaiming_data.filter(shift__name=shift_name)
    
    # Order by timestamp
    reclaiming_data = reclaiming_data.order_by('-timestamp')
    
    context = {
        'selected_date': selected_date,
        'areas': areas,
        'materials': materials,
        'shifts': shifts,
        'selected_area': selected_area,
        'selected_shift': selected_shift,
        'reclaiming_data': reclaiming_data,
    }
    
    return render(request, 'operations/reclaiming.html', context)

def crushing_view(request):
    """View for crushing data entry and display"""
    # Get filter parameters
    date_str = request.GET.get('date')
    area_id = request.GET.get('area')
    shift_name = request.GET.get('shift')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get all areas and materials for dropdowns
    areas = Area.objects.all()
    materials = Material.objects.all()
    
    # Get shifts for the selected date
    shifts = Shift.objects.filter(date=selected_date).order_by('name')
    if not shifts.exists():
        # Create shifts for this date if they don't exist
        for shift_name in ['A', 'B', 'C']:
            Shift.objects.create(name=shift_name, date=selected_date)
        shifts = Shift.objects.filter(date=selected_date).order_by('name')
    
    selected_area = None
    if area_id:
        selected_area = get_object_or_404(Area, id=area_id)
    
    selected_shift = shift_name
    
    # Handle form submission
    if request.method == 'POST':
        area_id = request.POST.get('area')
        shift_id = request.POST.get('shift')
        crusher = request.POST.get('crusher')
        notes = request.POST.get('notes')
        reported_by = request.POST.get('reported_by')
        
        # Get material names, IDs, and quantities as lists
        material_names = request.POST.getlist('material_names[]')
        material_ids = request.POST.getlist('material_ids[]')
        quantities = request.POST.getlist('quantities[]')
        
        # Validate form data
        if area_id and shift_id and crusher and reported_by and material_names and quantities:
            try:
                area = Area.objects.get(id=area_id)
                shift = Shift.objects.get(id=shift_id)
                
                # Create entries for each material
                success_count = 0
                for i in range(len(material_names)):
                    if material_names[i] and quantities[i]:
                        try:
                            # Try to get material by ID first
                            material = None
                            if material_ids[i]:
                                try:
                                    material = Material.objects.get(id=material_ids[i])
                                except Material.DoesNotExist:
                                    pass
                            
                            # If no material found by ID, try to find by name
                            if not material:
                                material_name = material_names[i].strip()
                                try:
                                    material = Material.objects.get(name__iexact=material_name)
                                except Material.DoesNotExist:
                                    # Create new material if it doesn't exist
                                    material = Material.objects.create(
                                        name=material_name,
                                        is_active=True
                                    )
                                    messages.info(request, f'Created new material: {material_name}')
                            
                            quantity = float(quantities[i])
                            
                            # Create new crushing entry
                            CrushingOperation.objects.create(
                                area=area,
                                shift=shift,
                                material=material,
                                quantity=quantity,
                                crusher=crusher,
                                notes=notes,
                                reported_by=reported_by
                            )
                            success_count += 1
                        except Exception as e:
                            messages.error(request, f'Error adding material {i+1}: {str(e)}')
                
                if success_count > 0:
                    messages.success(request, f'{success_count} crushing entries added successfully!')
                    return redirect('crushing')
            except Exception as e:
                messages.error(request, f'Error adding crushing entries: {str(e)}')
        else:
            messages.error(request, 'Please fill all required fields')
    
    # Get crushing data for display
    crushing_data = CrushingOperation.objects.filter(shift__date=selected_date)
    
    # Apply filters if provided
    if area_id:
        crushing_data = crushing_data.filter(area_id=area_id)
    
    if shift_name:
        crushing_data = crushing_data.filter(shift__name=shift_name)
    
    # Order by timestamp
    crushing_data = crushing_data.order_by('-timestamp')
    
    context = {
        'selected_date': selected_date,
        'areas': areas,
        'materials': materials,
        'shifts': shifts,
        'selected_area': selected_area,
        'selected_shift': selected_shift,
        'crushing_data': crushing_data,
    }
    
    return render(request, 'operations/crushing.html', context)

def receiving_view(request):
    """View for receiving data entry and display"""
    # Get filter parameters
    date_str = request.GET.get('date')
    area_id = request.GET.get('area')
    shift_name = request.GET.get('shift')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get all areas and materials for dropdowns
    areas = Area.objects.all()
    materials = Material.objects.all()
    
    # Get shifts for the selected date
    shifts = Shift.objects.filter(date=selected_date).order_by('name')
    if not shifts.exists():
        # Create shifts for this date if they don't exist
        for shift_name in ['A', 'B', 'C']:
            Shift.objects.create(name=shift_name, date=selected_date)
        shifts = Shift.objects.filter(date=selected_date).order_by('name')
    
    selected_area = None
    if area_id:
        selected_area = get_object_or_404(Area, id=area_id)
    
    selected_shift = shift_name
    
    # Handle form submission
    if request.method == 'POST':
        area_id = request.POST.get('area')
        shift_id = request.POST.get('shift')
        source = request.POST.get('source')  # Get the value from 'source' field in the form
        notes = request.POST.get('notes')
        reported_by = request.POST.get('reported_by')
        
        # Get material names, IDs, and quantities as lists
        material_names = request.POST.getlist('material_names[]')
        material_ids = request.POST.getlist('material_ids[]')
        quantities = request.POST.getlist('quantities[]')
        
        # Validate form data
        if area_id and shift_id and source and reported_by and material_names and quantities:
            try:
                area = Area.objects.get(id=area_id)
                shift = Shift.objects.get(id=shift_id)
                
                # Create entries for each material
                success_count = 0
                for i in range(len(material_names)):
                    if material_names[i] and quantities[i]:
                        try:
                            # Try to get material by ID first
                            material = None
                            if material_ids[i]:
                                try:
                                    material = Material.objects.get(id=material_ids[i])
                                except Material.DoesNotExist:
                                    pass
                            
                            # If no material found by ID, try to find by name
                            if not material:
                                material_name = material_names[i].strip()
                                try:
                                    material = Material.objects.get(name__iexact=material_name)
                                except Material.DoesNotExist:
                                    # Create new material if it doesn't exist
                                    material = Material.objects.create(
                                        name=material_name,
                                        is_active=True
                                    )
                                    messages.info(request, f'Created new material: {material_name}')
                            
                            quantity = float(quantities[i])
                            
                            # Create new receiving entry
                            ReceivingOperation.objects.create(
                                area=area,
                                shift=shift,
                                material=material,
                                quantity=quantity,
                                source=source,  # Use 'source' field instead of 'destination'
                                notes=notes,
                                reported_by=reported_by
                            )
                            success_count += 1
                        except Exception as e:
                            messages.error(request, f'Error adding material {i+1}: {str(e)}')
                
                if success_count > 0:
                    messages.success(request, f'{success_count} receiving entries added successfully!')
                    return redirect('receiving')
            except Exception as e:
                messages.error(request, f'Error adding receiving entries: {str(e)}')
        else:
            messages.error(request, 'Please fill all required fields')
    
    # Get receiving data for display
    receiving_data = ReceivingOperation.objects.filter(shift__date=selected_date)
    
    # Apply filters if provided
    if area_id:
        receiving_data = receiving_data.filter(area_id=area_id)
    
    if shift_name:
        receiving_data = receiving_data.filter(shift__name=shift_name)
    
    # Order by timestamp
    receiving_data = receiving_data.order_by('-timestamp')
    
    context = {
        'selected_date': selected_date,
        'areas': areas,
        'materials': materials,
        'shifts': shifts,
        'selected_area': selected_area,
        'selected_shift': selected_shift,
        'receiving_data': receiving_data,
    }
    
    return render(request, 'operations/receiving.html', context)

def daily_summary_view(request):
    """View for daily summary of operations"""
    # Get filter parameters
    date_str = request.GET.get('date')
    area_id = request.GET.get('area')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get all areas for filter dropdown
    areas = Area.objects.all()
    selected_area = None
    if area_id:
        selected_area = get_object_or_404(Area, id=area_id)
    
    # Build base querysets
    feeding_qs = FeedingOperation.objects.filter(shift__date=selected_date)
    reclaiming_qs = ReclaimingOperation.objects.filter(shift__date=selected_date)
    crushing_qs = CrushingOperation.objects.filter(shift__date=selected_date)
    receiving_qs = ReceivingOperation.objects.filter(shift__date=selected_date)
    
    # Apply area filter if selected
    if selected_area:
        feeding_qs = feeding_qs.filter(area=selected_area)
        reclaiming_qs = reclaiming_qs.filter(area=selected_area)
        crushing_qs = crushing_qs.filter(area=selected_area)
        receiving_qs = receiving_qs.filter(area=selected_area)
    
    # Calculate totals
    feeding_total = feeding_qs.aggregate(total=Sum('quantity'))['total'] or 0
    reclaiming_total = reclaiming_qs.aggregate(total=Sum('quantity'))['total'] or 0
    crushing_total = crushing_qs.aggregate(total=Sum('quantity'))['total'] or 0
    receiving_total = receiving_qs.aggregate(total=Sum('quantity'))['total'] or 0
    
    # Get material-wise summary
    materials = Material.objects.all()
    material_summary = []
    
    for material in materials:
        material_data = {
            'name': material.name,
            'feeding': feeding_qs.filter(material=material).aggregate(total=Sum('quantity'))['total'] or 0,
            'reclaiming': reclaiming_qs.filter(material=material).aggregate(total=Sum('quantity'))['total'] or 0,
            'crushing': crushing_qs.filter(material=material).aggregate(total=Sum('quantity'))['total'] or 0,
            'receiving': receiving_qs.filter(material=material).aggregate(total=Sum('quantity'))['total'] or 0,
        }
        material_data['total'] = (
            material_data['feeding'] + 
            material_data['reclaiming'] + 
            material_data['crushing'] + 
            material_data['receiving']
        )
        material_summary.append(material_data)
    
    # Get area-wise summary
    area_summary = []
    
    for area in areas:
        area_data = {
            'name': area.name,
            'feeding': feeding_qs.filter(area=area).aggregate(total=Sum('quantity'))['total'] or 0,
            'reclaiming': reclaiming_qs.filter(area=area).aggregate(total=Sum('quantity'))['total'] or 0,
            'crushing': crushing_qs.filter(area=area).aggregate(total=Sum('quantity'))['total'] or 0,
            'receiving': receiving_qs.filter(area=area).aggregate(total=Sum('quantity'))['total'] or 0,
        }
        area_data['total'] = (
            area_data['feeding'] + 
            area_data['reclaiming'] + 
            area_data['crushing'] + 
            area_data['receiving']
        )
        area_summary.append(area_data)
    
    # Get shift-wise summary
    shift_summary = {
        'A': {
            'feeding': feeding_qs.filter(shift__name='A').aggregate(total=Sum('quantity'))['total'] or 0,
            'reclaiming': reclaiming_qs.filter(shift__name='A').aggregate(total=Sum('quantity'))['total'] or 0,
            'crushing': crushing_qs.filter(shift__name='A').aggregate(total=Sum('quantity'))['total'] or 0,
            'receiving': receiving_qs.filter(shift__name='A').aggregate(total=Sum('quantity'))['total'] or 0,
        },
        'B': {
            'feeding': feeding_qs.filter(shift__name='B').aggregate(total=Sum('quantity'))['total'] or 0,
            'reclaiming': reclaiming_qs.filter(shift__name='B').aggregate(total=Sum('quantity'))['total'] or 0,
            'crushing': crushing_qs.filter(shift__name='B').aggregate(total=Sum('quantity'))['total'] or 0,
            'receiving': receiving_qs.filter(shift__name='B').aggregate(total=Sum('quantity'))['total'] or 0,
        },
        'C': {
            'feeding': feeding_qs.filter(shift__name='C').aggregate(total=Sum('quantity'))['total'] or 0,
            'reclaiming': reclaiming_qs.filter(shift__name='C').aggregate(total=Sum('quantity'))['total'] or 0,
            'crushing': crushing_qs.filter(shift__name='C').aggregate(total=Sum('quantity'))['total'] or 0,
            'receiving': receiving_qs.filter(shift__name='C').aggregate(total=Sum('quantity'))['total'] or 0,
        },
    }
    
    context = {
        'selected_date': selected_date,
        'areas': areas,
        'selected_area': selected_area,
        'feeding_total': feeding_total,
        'reclaiming_total': reclaiming_total,
        'crushing_total': crushing_total,
        'receiving_total': receiving_total,
        'material_summary': material_summary,
        'area_summary': area_summary,
        'shift_summary': shift_summary,
    }
    
    return render(request, 'operations/daily_summary.html', context)

@require_POST
def export_data(request):
    """API endpoint to export operations data as JSON, XLSX, or PDF"""
    date_str = request.POST.get('date')
    area_id = request.POST.get('area')
    shift_name = request.POST.get('shift')
    export_format = request.POST.get('format', 'json')  # Default to JSON if not specified
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Build filter
    filters = {'shift__date': selected_date}
    
    if area_id:
        filters['area_id'] = area_id
    
    if shift_name:
        filters['shift__name'] = shift_name
    
    # Get data
    feeding_data = []
    for item in FeedingOperation.objects.filter(**filters):
        feeding_data.append({
            'id': item.id,
            'area': item.area.name,
            'shift': item.shift.get_name_display(),
            'material': item.material.name,
            'quantity': float(item.quantity),
            'destination': item.destination,
            'reported_by': item.reported_by,
            'timestamp': item.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'notes': item.notes,
        })
    
    reclaiming_data = []
    for item in ReclaimingOperation.objects.filter(**filters):
        reclaiming_data.append({
            'id': item.id,
            'area': item.area.name,
            'shift': item.shift.get_name_display(),
            'material': item.material.name,
            'quantity': float(item.quantity),
            'source': item.source,
            'reported_by': item.reported_by,
            'timestamp': item.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'notes': item.notes,
        })
    
    crushing_data = []
    for item in CrushingOperation.objects.filter(**filters):
        crushing_data.append({
            'id': item.id,
            'area': item.area.name,
            'shift': item.shift.get_name_display(),
            'material': item.material.name,
            'quantity': float(item.quantity),
            'crusher': item.crusher,
            'reported_by': item.reported_by,
            'timestamp': item.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'notes': item.notes,
        })
    
    receiving_data = []
    for item in ReceivingOperation.objects.filter(**filters):
        receiving_data.append({
            'id': item.id,
            'area': item.area.name,
            'shift': item.shift.get_name_display(),
            'material': item.material.name,
            'quantity': float(item.quantity),
            'source': item.source,
            'reported_by': item.reported_by,
            'timestamp': item.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            'notes': item.notes,
        })
    
    # Prepare data for all formats
    data = {
        'date': selected_date.strftime('%Y-%m-%d'),
        'feeding': feeding_data,
        'reclaiming': reclaiming_data,
        'crushing': crushing_data,
        'receiving': receiving_data,
    }
    
    # Export based on requested format
    if export_format == 'xlsx':
        return export_as_xlsx(data, selected_date)
    elif export_format == 'pdf':
        return export_as_pdf(data, selected_date)
    else:  # Default to JSON
        return JsonResponse(data)

def export_as_xlsx(data, selected_date):
    """Export data as Excel file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    import io
    
    try:
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Summary"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        centered_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        
        # Add title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"Daily Operations Summary - {selected_date.strftime('%B %d, %Y')}"
        title_cell.font = Font(name='Arial', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Add section headers and data for each operation type
        operation_types = [
            ('Feeding Operations', data['feeding']),
            ('Reclaiming Operations', data['reclaiming']),
            ('Crushing Operations', data['crushing']),
            ('Receiving Operations', data['receiving'])
        ]
        
        row = 3
        for section_title, section_data in operation_types:
            # Add section header
            ws.merge_cells(f'A{row}:H{row}')
            section_header = ws[f'A{row}']
            section_header.value = section_title
            section_header.font = Font(name='Arial', size=14, bold=True)
            section_header.alignment = Alignment(horizontal='left', vertical='center')
            ws.row_dimensions[row].height = 25
            row += 1
            
            # Skip if no data
            if not section_data:
                ws.merge_cells(f'A{row}:H{row}')
                no_data_cell = ws[f'A{row}']
                no_data_cell.value = "No data available"
                no_data_cell.alignment = Alignment(horizontal='center', vertical='center')
                row += 2
                continue
            
            # Add table headers
            headers = ['ID', 'Area', 'Shift', 'Material', 'Quantity (tons)', 'Location', 'Reported By', 'Timestamp']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered_alignment
                cell.border = border
                
            ws.row_dimensions[row].height = 20
            row += 1
            
            # Add data rows
            for item in section_data:
                ws.cell(row=row, column=1).value = item['id']
                ws.cell(row=row, column=2).value = item['area']
                ws.cell(row=row, column=3).value = item['shift']
                ws.cell(row=row, column=4).value = item['material']
                ws.cell(row=row, column=5).value = item['quantity']
                
                # Location field varies by operation type
                if 'destination' in item:
                    ws.cell(row=row, column=6).value = item['destination']
                elif 'source' in item:
                    ws.cell(row=row, column=6).value = item['source']
                elif 'crusher' in item:
                    ws.cell(row=row, column=6).value = item['crusher']
                    
                ws.cell(row=row, column=7).value = item['reported_by']
                ws.cell(row=row, column=8).value = item['timestamp']
                
                # Apply borders to all cells in the row
                for col_idx in range(1, 9):
                    ws.cell(row=row, column=col_idx).border = border
                    
                row += 1
            
            # Add empty row between sections
            row += 1
        
        # Auto-adjust column widths
        for col_idx in range(1, 9):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
        
        # Special width adjustments
        ws.column_dimensions['D'].width = 20  # Material column
        ws.column_dimensions['E'].width = 15  # Quantity column
        ws.column_dimensions['F'].width = 20  # Location column
        ws.column_dimensions['H'].width = 20  # Timestamp column
        
        # Create a buffer to save the workbook
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=daily_summary_{selected_date.strftime("%Y-%m-%d")}.xlsx'
        
        return response
        
    except Exception as e:
        import traceback
        print(f"Excel Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        # If there's an error, return a plain text response with the error
        return HttpResponse(f"Error generating Excel file: {str(e)}", content_type='text/plain')

def export_as_pdf(data, selected_date):
    """Export data as PDF file"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from django.http import HttpResponse
    import io
    
    # Create a buffer to receive PDF data
    buffer = io.BytesIO()
    
    try:
        # Create the PDF document using the buffer
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.alignment = 1  # Center alignment
        normal_style = styles['Normal']
        
        # Create elements to add to the PDF
        elements = []
        
        # Add title
        title = Paragraph(f"Daily Operations Summary - {selected_date.strftime('%B %d, %Y')}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Add section for each operation type
        operation_types = [
            ('Feeding Operations', data['feeding']),
            ('Reclaiming Operations', data['reclaiming']),
            ('Crushing Operations', data['crushing']),
            ('Receiving Operations', data['receiving'])
        ]
        
        for section_title, section_data in operation_types:
            # Add section header
            section_header = Paragraph(section_title, styles['Heading2'])
            elements.append(section_header)
            elements.append(Spacer(1, 10))
            
            # Skip if no data
            if not section_data:
                no_data = Paragraph("No data available", normal_style)
                elements.append(no_data)
                elements.append(Spacer(1, 20))
                continue
            
            # Prepare table data
            headers = ['ID', 'Area', 'Shift', 'Material', 'Quantity', 'Location', 'Reported By', 'Timestamp']
            table_data = [headers]
            
            for item in section_data:
                # Determine location field based on operation type
                if 'destination' in item:
                    location = item['destination']
                elif 'source' in item:
                    location = item['source']
                elif 'crusher' in item:
                    location = item['crusher']
                else:
                    location = ''
                
                row = [
                    str(item['id']),
                    item['area'],
                    item['shift'],
                    item['material'],
                    f"{item['quantity']:.2f}",
                    location,
                    item['reported_by'],
                    item['timestamp']
                ]
                table_data.append(row)
            
            # Create table
            table = Table(table_data)
            
            # Apply table styles
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                
                # Data row styling
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # ID column centered
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Quantity column right-aligned
                
                # Grid styling
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                
                # Padding
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ]))
            
            # Add table to elements
            elements.append(table)
            elements.append(Spacer(1, 20))
        
        # Build the PDF document
        doc.build(elements)
        
        # Get the value of the buffer
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Create HTTP response with PDF file
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=daily_summary_{selected_date.strftime("%Y-%m-%d")}.pdf'
        response.write(pdf_data)
        
        return response
        
    except Exception as e:
        import traceback
        print(f"PDF Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        # If there's an error, return a plain text response with the error
        return HttpResponse(f"Error generating PDF: {str(e)}", content_type='text/plain')
