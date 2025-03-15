from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Count, Q
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from .models import RakeEntry, RakeNotification
from operations.models import Area, Shift, Material
from datetime import datetime, timedelta, time
import io
import sys

def rake_dashboard(request):
    """Main rake handling dashboard view"""
    # Get filter parameters
    date_str = request.GET.get('date')
    status = request.GET.get('status')
    wagon_tippler = request.GET.get('wagon_tippler')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Build base querysets
    start_date = datetime.combine(selected_date, time.min)
    end_date = datetime.combine(selected_date, time.max)
    rake_qs = RakeEntry.objects.filter(arrival_time__range=(start_date, end_date))
    
    # Apply status filter if selected
    if status:
        rake_qs = rake_qs.filter(status=status)
        
    # Apply wagon tippler filter if selected
    if wagon_tippler:
        rake_qs = rake_qs.filter(wagon_tippler=wagon_tippler)
    
    # Get summary data
    completed_count = rake_qs.filter(status='COMPLETED').count()
    in_progress_count = rake_qs.filter(status='IN_PROGRESS').count()
    pending_count = rake_qs.filter(status='PENDING').count()
    
    # Get unread notifications
    notifications = RakeNotification.objects.filter(is_read=False).order_by('-timestamp')[:5]
    
    # Get wagon tipplers
    wagon_tipplers = [choice for choice in RakeEntry.WAGON_TIPPLER_CHOICES]
    
    # Get materials
    materials = Material.objects.all()
    
    context = {
        'selected_date': selected_date,
        'status': status,
        'rake_data': rake_qs,
        'completed_count': completed_count,
        'in_progress_count': in_progress_count,
        'pending_count': pending_count,
        'notifications': notifications,
        'wagon_tipplers': wagon_tipplers,
        'selected_wagon_tippler': wagon_tippler,
        'materials': materials,
    }
    
    return render(request, 'rake_handling/dashboard.html', context)

def rake_entry_view(request):
    """View for rake entry form and data"""
    print("DEBUGGING: Rake Entry View accessed")
    print(f"DEBUGGING: Request path: {request.path}")
    
    # Get the selected date from the request or use today's date
    selected_date = request.GET.get('date', timezone.now().date())
    if isinstance(selected_date, str):
        try:
            selected_date = datetime.strptime(selected_date, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    
    # Get the selected wagon tippler from the request
    selected_wagon_tippler = request.GET.get('wagon_tippler', '')
    
    # Get all materials for the dropdown
    materials = Material.objects.all()
    
    # Define wagon tippler choices
    wagon_tipplers = RakeEntry.WAGON_TIPPLER_CHOICES
    print(f"DEBUGGING: Wagon tipplers: {wagon_tipplers}")
    
    # Process form submission
    if request.method == 'POST':
        # Get form data
        wagon_tippler = request.POST.get('wagon_tippler')
        material_id = request.POST.get('material')
        material_name = request.POST.get('material_name')
        rake_id = request.POST.get('rake_id')
        quantity = request.POST.get('quantity', 0)
        status = request.POST.get('status')
        reported_by = request.POST.get('reported_by')
        notes = request.POST.get('notes')
        
        # Validate required fields
        if not wagon_tippler or not rake_id or not status or not reported_by:
            messages.error(request, 'Please fill in all required fields.')
            return redirect('rake_entry')
        
        # Handle material (either selected from dropdown or manually entered)
        if material_name:
            # Try to find existing material or create new one
            material, created = Material.objects.get_or_create(name=material_name)
            if created:
                messages.info(request, f'New material "{material_name}" has been created.')
        elif material_id:
            try:
                material = Material.objects.get(id=material_id)
            except Material.DoesNotExist:
                messages.error(request, 'Selected material does not exist.')
                return redirect('rake_entry')
        else:
            messages.error(request, 'Please provide a material.')
            return redirect('rake_entry')
        
        # Create new rake entry
        rake_entry = RakeEntry.objects.create(
            wagon_tippler=wagon_tippler,
            material=material,
            rake_id=rake_id,
            quantity=quantity if quantity else 0,
            status=status,
            reported_by=reported_by,
            notes=notes
        )
        
        # Create notification for new rake entry
        RakeNotification.objects.create(
            rake_entry=rake_entry,
            message=f"New rake {rake_id} with {material.name} registered at {rake_entry.get_wagon_tippler_display()}",
            status=status
        )
        
        messages.success(request, 'Rake entry added successfully.')
        return redirect('rake_entry')
    
    # Get rake data for the selected date
    start_date = datetime.combine(selected_date, time.min)
    end_date = datetime.combine(selected_date, time.max)
    
    rake_query = RakeEntry.objects.filter(arrival_time__range=(start_date, end_date))
    
    # Filter by wagon tippler if selected
    if selected_wagon_tippler:
        rake_query = rake_query.filter(wagon_tippler=selected_wagon_tippler)
    
    rake_data = rake_query.order_by('-arrival_time')
    
    # Calculate material summary
    material_summary = {}
    for rake in rake_data:
        material_name = rake.material.name
        if material_name not in material_summary:
            material_summary[material_name] = {
                'count': 0,
                'total_quantity': 0
            }
        material_summary[material_name]['count'] += 1
        material_summary[material_name]['total_quantity'] += rake.quantity
    
    # Prepare context for template
    context = {
        'selected_date': selected_date,
        'materials': materials,
        'status': RakeEntry.STATUS_CHOICES,
        'rake_data': rake_data,
        'wagon_tipplers': wagon_tipplers,
        'selected_wagon_tippler': selected_wagon_tippler,
        'material_summary': material_summary,
    }
    
    print("DEBUGGING: Using template: rake_handling/rake_entry.html")
    print(f"DEBUGGING: Context keys: {context.keys()}")
    
    return render(request, 'rake_handling/rake_entry.html', context)

def notifications_view(request):
    """View for rake notifications"""
    # Get all notifications
    notifications = RakeNotification.objects.all().order_by('-timestamp')
    
    # Get unread count
    unread_count = notifications.filter(is_read=False).count()
    
    context = {
        'notifications': notifications,
        'unread_count': unread_count,
    }
    
    return render(request, 'rake_handling/notifications.html', context)

@require_POST
def update_rake_status(request, rake_id):
    """API endpoint to update rake status"""
    rake = get_object_or_404(RakeEntry, id=rake_id)
    status = request.POST.get('status')
    notes = request.POST.get('notes', '')
    
    if status in ['COMPLETED', 'IN_PROGRESS', 'PENDING']:
        rake.update_status(status, notes)
        return JsonResponse({'status': 'success', 'message': f'Status updated to {status}'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid status'}, status=400)

@require_POST
def mark_notification_read(request, notification_id):
    """API endpoint to mark notification as read"""
    notification = get_object_or_404(RakeNotification, id=notification_id)
    notification.is_read = True
    notification.save()
    return JsonResponse({'status': 'success', 'message': 'Notification marked as read'})

@require_POST
def export_rake_data(request):
    """API endpoint to export rake data as JSON, XLSX, or PDF"""
    date_str = request.POST.get('date')
    wagon_tippler = request.POST.get('wagon_tippler')
    status = request.POST.get('status')
    export_format = request.POST.get('format', 'json')  # Default to JSON if not specified
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Build filter
    start_date = datetime.combine(selected_date, time.min)
    end_date = datetime.combine(selected_date, time.max)
    filters = {'arrival_time__range': (start_date, end_date)}
    
    if status:
        filters['status'] = status
    
    if wagon_tippler:
        filters['wagon_tippler'] = wagon_tippler
    
    # Get data
    queryset = RakeEntry.objects.filter(**filters)
    data = []
    
    for item in queryset:
        data.append({
            'id': item.id,
            'rake_id': item.rake_id,
            'material': item.material.name,
            'quantity': float(item.quantity),
            'status': item.get_status_display(),
            'wagon_tippler': item.get_wagon_tippler_display(),
            'arrival_time': item.arrival_time.strftime('%Y-%m-%d %H:%M:%S'),
            'completion_time': item.completion_time.strftime('%Y-%m-%d %H:%M:%S') if item.completion_time else None,
            'notes': item.notes,
            'reported_by': item.reported_by,
        })
    
    # Export based on requested format
    if export_format == 'xlsx':
        return export_rake_as_xlsx(data, selected_date)
    elif export_format == 'pdf':
        return export_rake_as_pdf(data, selected_date)
    else:  # Default to JSON
        return JsonResponse({'data': data})

def export_rake_as_xlsx(data, selected_date):
    """Export rake data as Excel file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    import io
    
    try:
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Rake Entries"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        centered_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        
        # Add title
        ws.merge_cells('A1:J1')
        title_cell = ws['A1']
        title_cell.value = f"Rake Entries Summary - {selected_date.strftime('%B %d, %Y')}"
        title_cell.font = Font(name='Arial', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Add table headers
        headers = ['ID', 'Rake ID', 'Wagon Tippler', 'Material', 'Quantity (tons)', 'Status', 'Arrival Time', 'Completion Time', 'Reported By']
        row = 3
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = border
            
        ws.row_dimensions[row].height = 20
        row += 1
        
        # Add data rows
        for item in data:
            ws.cell(row=row, column=1).value = item['id']
            ws.cell(row=row, column=2).value = item['rake_id']
            ws.cell(row=row, column=3).value = item['wagon_tippler']
            ws.cell(row=row, column=4).value = item['material']
            ws.cell(row=row, column=5).value = item['quantity']
            ws.cell(row=row, column=6).value = item['status']
            ws.cell(row=row, column=7).value = item['arrival_time']
            ws.cell(row=row, column=8).value = item['completion_time'] if item['completion_time'] else 'N/A'
            ws.cell(row=row, column=9).value = item['reported_by']
            
            # Apply borders to all cells in the row
            for col_idx in range(1, 10):
                ws.cell(row=row, column=col_idx).border = border
                
            row += 1
        
        # Add summary section
        row += 2
        summary_title = ws.cell(row=row, column=1)
        summary_title.value = "Summary by Material"
        summary_title.font = Font(name='Arial', size=14, bold=True)
        ws.merge_cells(f'A{row}:J{row}')
        summary_title.alignment = Alignment(horizontal='center', vertical='center')
        row += 1
        
        # Add summary headers
        summary_headers = ['Material', 'Number of Rakes', 'Total Quantity (tons)']
        for col_idx, header in enumerate(summary_headers, 1):
            cell = ws.cell(row=row, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = centered_alignment
            cell.border = border
        row += 1
        
        # Group data by material
        material_summary = {}
        for item in data:
            material = item['material']
            if material not in material_summary:
                material_summary[material] = {
                    'count': 0,
                    'total_quantity': 0
                }
            material_summary[material]['count'] += 1
            material_summary[material]['total_quantity'] += item['quantity']
        
        # Add summary data
        for material, summary in material_summary.items():
            ws.cell(row=row, column=1).value = material
            ws.cell(row=row, column=2).value = summary['count']
            ws.cell(row=row, column=3).value = summary['total_quantity']
            
            # Apply borders to all cells in the row
            for col_idx in range(1, 4):
                ws.cell(row=row, column=col_idx).border = border
            
            row += 1
        
        # Auto-adjust column widths
        for col_idx in range(1, 10):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
        
        # Special width adjustments
        ws.column_dimensions['B'].width = 15  # Rake ID column
        ws.column_dimensions['C'].width = 20  # Wagon Tippler column
        ws.column_dimensions['D'].width = 20  # Material column
        ws.column_dimensions['G'].width = 20  # Arrival Time column
        ws.column_dimensions['H'].width = 20  # Completion Time column
        
        # Create a buffer to save the workbook
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=rake_summary_{selected_date.strftime("%Y-%m-%d")}.xlsx'
        
        return response
        
    except Exception as e:
        import traceback
        print(f"Excel Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        # If there's an error, return a plain text response with the error
        return HttpResponse(f"Error generating Excel file: {str(e)}", content_type='text/plain')

def export_rake_as_pdf(data, selected_date):
    """Export rake data as PDF file"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from django.http import HttpResponse
    import io
    
    # Create a buffer to receive PDF data
    buffer = io.BytesIO()
    
    try:
        # Create the PDF document using the buffer
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.alignment = 1  # Center alignment
        normal_style = styles['Normal']
        
        # Create elements to add to the PDF
        elements = []
        
        # Add title
        title = Paragraph(f"Rake Entries Summary - {selected_date.strftime('%B %d, %Y')}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 20))
        
        # Skip if no data
        if not data:
            no_data = Paragraph("No data available", normal_style)
            elements.append(no_data)
        else:
            # Prepare table data
            headers = ['ID', 'Rake ID', 'Wagon Tippler', 'Material', 'Quantity', 'Status', 'Arrival Time', 'Completion Time']
            table_data = [headers]
            
            for item in data:
                row = [
                    str(item['id']),
                    item['rake_id'],
                    item['wagon_tippler'],
                    item['material'],
                    f"{item['quantity']:.2f}",
                    item['status'],
                    item['arrival_time'],
                    item['completion_time'] if item['completion_time'] else 'N/A'
                ]
                table_data.append(row)
            
            # Create table
            table = Table(table_data)
            
            # Apply table styles
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                
                # Data row styling
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # ID column centered
                ('ALIGN', (4, 1), (4, -1), 'RIGHT'),   # Quantity column right-aligned
                
                # Grid styling
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                
                # Padding
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ]))
            
            # Add table to elements
            elements.append(table)
            
            # Add summary section
            elements.append(Spacer(1, 20))
            summary_title = Paragraph("Summary by Material", styles['Heading2'])
            elements.append(summary_title)
            elements.append(Spacer(1, 10))
            
            # Group data by material
            material_summary = {}
            for item in data:
                material = item['material']
                if material not in material_summary:
                    material_summary[material] = {
                        'count': 0,
                        'total_quantity': 0
                    }
                material_summary[material]['count'] += 1
                material_summary[material]['total_quantity'] += item['quantity']
            
            # Create summary table
            summary_headers = ['Material', 'Number of Rakes', 'Total Quantity (tons)']
            summary_data = [summary_headers]
            
            for material, summary in material_summary.items():
                row = [
                    material,
                    str(summary['count']),
                    f"{summary['total_quantity']:.2f}"
                ]
                summary_data.append(row)
            
            summary_table = Table(summary_data)
            summary_table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                
                # Data row styling
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (1, 1), (1, -1), 'CENTER'),  # Count column centered
                ('ALIGN', (2, 1), (2, -1), 'RIGHT'),   # Quantity column right-aligned
                
                # Grid styling
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                
                # Padding
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ]))
            
            elements.append(summary_table)
        
        # Build the PDF document
        doc.build(elements)
        
        # Get the value of the buffer
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Create HTTP response with PDF file
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=rake_summary_{selected_date.strftime("%Y-%m-%d")}.pdf'
        response.write(pdf_data)
        
        return response
        
    except Exception as e:
        import traceback
        print(f"PDF Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        # If there's an error, return a plain text response with the error
        return HttpResponse(f"Error generating PDF: {str(e)}", content_type='text/plain') 