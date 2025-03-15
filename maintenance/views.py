from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db.models import Count, Q
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from .models import IssueType, MaintenanceIssue, PlannedMaintenance
from operations.models import Area, Shift
from datetime import datetime, timedelta

def maintenance_dashboard(request):
    """Main maintenance dashboard view"""
    # Get filter parameters
    date_str = request.GET.get('date')
    area_id = request.GET.get('area')
    category = request.GET.get('category')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get all areas for filter dropdown
    areas = Area.objects.all()
    selected_area = None
    if area_id:
        selected_area = get_object_or_404(Area, id=area_id)
    
    # Build base querysets
    issues_qs = MaintenanceIssue.objects.filter(shift__date=selected_date)
    planned_qs = PlannedMaintenance.objects.filter(planned_date=selected_date)
    
    # Apply area filter if selected
    if selected_area:
        issues_qs = issues_qs.filter(area=selected_area)
        planned_qs = planned_qs.filter(area=selected_area)
    
    # Apply category filter if selected
    if category:
        issues_qs = issues_qs.filter(issue_type__category=category)
        planned_qs = planned_qs.filter(category=category)
    
    # Get summary data
    electrical_issues_count = issues_qs.filter(issue_type__category='ELECTRICAL').count()
    mechanical_issues_count = issues_qs.filter(issue_type__category='MECHANICAL').count()
    
    resolved_issues_count = issues_qs.filter(resolved=True).count()
    unresolved_issues_count = issues_qs.filter(resolved=False).count()
    
    planned_completed_count = planned_qs.filter(status='COMPLETED').count()
    planned_in_progress_count = planned_qs.filter(status='IN_PROGRESS').count()
    planned_pending_count = planned_qs.filter(status='PENDING').count()
    
    # Get shift-wise data
    shifts = Shift.objects.filter(date=selected_date).order_by('name')
    
    context = {
        'selected_date': selected_date,
        'areas': areas,
        'selected_area': selected_area,
        'category': category,
        'shifts': shifts,
        'issues_data': issues_qs,
        'planned_data': planned_qs,
        'electrical_issues_count': electrical_issues_count,
        'mechanical_issues_count': mechanical_issues_count,
        'resolved_issues_count': resolved_issues_count,
        'unresolved_issues_count': unresolved_issues_count,
        'planned_completed_count': planned_completed_count,
        'planned_in_progress_count': planned_in_progress_count,
        'planned_pending_count': planned_pending_count,
    }
    
    return render(request, 'maintenance/dashboard.html', context)

def electrical_issues_view(request):
    """View for electrical issues data entry and display"""
    # Get filter parameters
    date_str = request.GET.get('date')
    area_id = request.GET.get('area')
    shift_name = request.GET.get('shift')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get all areas and issue types for dropdowns
    areas = Area.objects.all()
    issue_types = IssueType.objects.filter(category='ELECTRICAL')
    
    # Get shifts for the selected date
    shifts = Shift.objects.filter(date=selected_date).order_by('name')
    if not shifts.exists():
        # Create shifts for this date if they don't exist
        for shift_name in ['A', 'B', 'C']:
            Shift.objects.create(name=shift_name, date=selected_date)
        shifts = Shift.objects.filter(date=selected_date).order_by('name')
    
    # Handle form submission
    if request.method == 'POST':
        area_id = request.POST.get('area')
        shift_id = request.POST.get('shift')
        issue_type_id = request.POST.get('issue_type')
        description = request.POST.get('description')
        reported_by = request.POST.get('reported_by')
        
        # Validate form data
        if all([area_id, shift_id, issue_type_id, description, reported_by]):
            try:
                area = Area.objects.get(id=area_id)
                shift = Shift.objects.get(id=shift_id)
                issue_type = IssueType.objects.get(id=issue_type_id)
                
                # Create new issue entry
                MaintenanceIssue.objects.create(
                    area=area,
                    shift=shift,
                    issue_type=issue_type,
                    description=description,
                    reported_by=reported_by
                )
                
                messages.success(request, 'Electrical issue added successfully!')
                return redirect('electrical_issues')
            except Exception as e:
                messages.error(request, f'Error adding electrical issue: {str(e)}')
        else:
            messages.error(request, 'Please fill all required fields')
    
    # Get electrical issues data for display
    issues_data = MaintenanceIssue.objects.filter(
        shift__date=selected_date,
        issue_type__category='ELECTRICAL'
    )
    
    # Apply filters if provided
    if area_id:
        issues_data = issues_data.filter(area_id=area_id)
    
    if shift_name:
        issues_data = issues_data.filter(shift__name=shift_name)
    
    # Order by timestamp
    issues_data = issues_data.order_by('-timestamp')
    
    context = {
        'selected_date': selected_date,
        'areas': areas,
        'issue_types': issue_types,
        'shifts': shifts,
        'issues_data': issues_data,
    }
    
    return render(request, 'maintenance/electrical_issues.html', context)

def mechanical_issues_view(request):
    """View for mechanical issues data entry and display"""
    # Similar structure to electrical_issues_view
    # Get filter parameters
    date_str = request.GET.get('date')
    area_id = request.GET.get('area')
    shift_name = request.GET.get('shift')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get all areas and issue types for dropdowns
    areas = Area.objects.all()
    issue_types = IssueType.objects.filter(category='MECHANICAL')
    
    # Get shifts for the selected date
    shifts = Shift.objects.filter(date=selected_date).order_by('name')
    if not shifts.exists():
        # Create shifts for this date if they don't exist
        for shift_name in ['A', 'B', 'C']:
            Shift.objects.create(name=shift_name, date=selected_date)
        shifts = Shift.objects.filter(date=selected_date).order_by('name')
    
    # Handle form submission
    if request.method == 'POST':
        area_id = request.POST.get('area')
        shift_id = request.POST.get('shift')
        issue_type_id = request.POST.get('issue_type')
        description = request.POST.get('description')
        reported_by = request.POST.get('reported_by')
        
        # Validate form data
        if all([area_id, shift_id, issue_type_id, description, reported_by]):
            try:
                area = Area.objects.get(id=area_id)
                shift = Shift.objects.get(id=shift_id)
                issue_type = IssueType.objects.get(id=issue_type_id)
                
                # Create new issue entry
                MaintenanceIssue.objects.create(
                    area=area,
                    shift=shift,
                    issue_type=issue_type,
                    description=description,
                    reported_by=reported_by
                )
                
                messages.success(request, 'Mechanical issue added successfully!')
                return redirect('mechanical_issues')
            except Exception as e:
                messages.error(request, f'Error adding mechanical issue: {str(e)}')
        else:
            messages.error(request, 'Please fill all required fields')
    
    # Get mechanical issues data for display
    issues_data = MaintenanceIssue.objects.filter(
        shift__date=selected_date,
        issue_type__category='MECHANICAL'
    )
    
    # Apply filters if provided
    if area_id:
        issues_data = issues_data.filter(area_id=area_id)
    
    if shift_name:
        issues_data = issues_data.filter(shift__name=shift_name)
    
    # Order by timestamp
    issues_data = issues_data.order_by('-timestamp')
    
    context = {
        'selected_date': selected_date,
        'areas': areas,
        'issue_types': issue_types,
        'shifts': shifts,
        'issues_data': issues_data,
    }
    
    return render(request, 'maintenance/mechanical_issues.html', context)

def planned_maintenance_view(request):
    """View for planned maintenance activities data entry and display"""
    # Get filter parameters
    date_str = request.GET.get('date')
    area_id = request.GET.get('area')
    category = request.GET.get('category')
    status = request.GET.get('status')
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get all areas for dropdowns
    areas = Area.objects.all()
    
    # Handle form submission
    if request.method == 'POST':
        area_id = request.POST.get('area')
        title = request.POST.get('title')
        description = request.POST.get('description')
        category = request.POST.get('category')
        planned_date = request.POST.get('planned_date')
        assigned_to = request.POST.get('assigned_to')
        
        # Validate form data
        if all([area_id, title, description, category, planned_date, assigned_to]):
            try:
                area = Area.objects.get(id=area_id)
                planned_date = datetime.strptime(planned_date, '%Y-%m-%d').date()
                
                # Create new planned maintenance entry
                PlannedMaintenance.objects.create(
                    area=area,
                    title=title,
                    description=description,
                    category=category,
                    planned_date=planned_date,
                    assigned_to=assigned_to
                )
                
                messages.success(request, 'Planned maintenance activity added successfully!')
                return redirect('planned_maintenance')
            except Exception as e:
                messages.error(request, f'Error adding planned maintenance: {str(e)}')
        else:
            messages.error(request, 'Please fill all required fields')
    
    # Get planned maintenance data for display
    planned_data = PlannedMaintenance.objects.filter(planned_date=selected_date)
    
    # Apply filters if provided
    if area_id:
        planned_data = planned_data.filter(area_id=area_id)
    
    if category:
        planned_data = planned_data.filter(category=category)
    
    if status:
        planned_data = planned_data.filter(status=status)
    
    # Order by planned_date
    planned_data = planned_data.order_by('planned_date')
    
    context = {
        'selected_date': selected_date,
        'areas': areas,
        'category': category,
        'status': status,
        'planned_data': planned_data,
    }
    
    return render(request, 'maintenance/planned_maintenance.html', context)

@require_POST
def update_issue_status(request, issue_id):
    """API endpoint to update issue status"""
    issue = get_object_or_404(MaintenanceIssue, id=issue_id)
    resolved = request.POST.get('resolved') == 'true'
    resolution_notes = request.POST.get('resolution_notes', '')
    
    if resolved:
        issue.resolve(resolution_notes)
        return JsonResponse({'status': 'success', 'message': 'Issue marked as resolved'})
    else:
        issue.resolved = False
        issue.resolution_notes = None
        issue.resolution_timestamp = None
        issue.save()
        return JsonResponse({'status': 'success', 'message': 'Issue marked as unresolved'})

@require_POST
def update_planned_status(request, planned_id):
    """API endpoint to update planned maintenance status"""
    planned = get_object_or_404(PlannedMaintenance, id=planned_id)
    status = request.POST.get('status')
    notes = request.POST.get('notes', '')
    
    if status in ['COMPLETED', 'IN_PROGRESS', 'PENDING']:
        planned.update_status(status, notes)
        return JsonResponse({'status': 'success', 'message': f'Status updated to {status}'})
    else:
        return JsonResponse({'status': 'error', 'message': 'Invalid status'}, status=400)

@require_POST
def export_maintenance_data(request):
    """API endpoint to export maintenance data as JSON, XLSX, or PDF"""
    date_str = request.POST.get('date')
    area_id = request.POST.get('area')
    data_type = request.POST.get('data_type')
    export_format = request.POST.get('format', 'json')  # Default to JSON if not specified
    
    # Default to today if no date provided
    if date_str:
        try:
            selected_date = datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Filter by area if provided
    area_filter = {}
    if area_id:
        area_filter['area_id'] = area_id
    
    # Get data based on type
    data = []
    title = "Maintenance Summary"
    
    if data_type == 'electrical_issues':
        queryset = MaintenanceIssue.objects.filter(
            shift__date=selected_date, 
            issue_type__category='ELECTRICAL',
            **area_filter
        )
        title = "Electrical Issues Summary"
        for item in queryset:
            data.append({
                'id': item.id,
                'area': item.area.name,
                'shift': item.shift.get_name_display(),
                'issue_type': item.issue_type.name,
                'description': item.description,
                'reported_by': item.reported_by,
                'resolved': item.resolved,
                'resolution_notes': item.resolution_notes,
                'timestamp': item.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            })
    elif data_type == 'mechanical_issues':
        queryset = MaintenanceIssue.objects.filter(
            shift__date=selected_date, 
            issue_type__category='MECHANICAL',
            **area_filter
        )
        title = "Mechanical Issues Summary"
        for item in queryset:
            data.append({
                'id': item.id,
                'area': item.area.name,
                'shift': item.shift.get_name_display(),
                'issue_type': item.issue_type.name,
                'description': item.description,
                'reported_by': item.reported_by,
                'resolved': item.resolved,
                'resolution_notes': item.resolution_notes,
                'timestamp': item.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            })
    elif data_type == 'planned_maintenance':
        queryset = PlannedMaintenance.objects.filter(
            scheduled_date=selected_date,
            **area_filter
        )
        title = "Planned Maintenance Summary"
        for item in queryset:
            data.append({
                'id': item.id,
                'area': item.area.name,
                'maintenance_type': item.get_maintenance_type_display(),
                'description': item.description,
                'scheduled_date': item.scheduled_date.strftime('%Y-%m-%d'),
                'scheduled_by': item.scheduled_by,
                'completed': item.completed,
                'completion_notes': item.completion_notes,
                'timestamp': item.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
            })
    
    # Export based on requested format
    if export_format == 'xlsx':
        return export_maintenance_as_xlsx(data, selected_date, data_type, title)
    elif export_format == 'pdf':
        return export_maintenance_as_pdf(data, selected_date, data_type, title)
    else:  # Default to JSON
        return JsonResponse({'data': data, 'title': title})

def export_maintenance_as_xlsx(data, selected_date, data_type, title):
    """Export maintenance data as Excel file"""
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    import io
    
    try:
        # Create a new workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Maintenance Data"
        
        # Define styles
        header_font = Font(name='Arial', size=12, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        centered_alignment = Alignment(horizontal='center', vertical='center')
        border = Border(
            left=Side(border_style='thin', color='000000'),
            right=Side(border_style='thin', color='000000'),
            top=Side(border_style='thin', color='000000'),
            bottom=Side(border_style='thin', color='000000')
        )
        
        # Add title
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = f"{title} - {selected_date.strftime('%B %d, %Y')}"
        title_cell.font = Font(name='Arial', size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30
        
        # Skip if no data
        if not data:
            ws.merge_cells('A3:I3')
            no_data_cell = ws['A3']
            no_data_cell.value = "No data available"
            no_data_cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            # Define headers based on data type
            if data_type == 'planned_maintenance':
                headers = ['ID', 'Area', 'Type', 'Description', 'Scheduled Date', 'Scheduled By', 'Completed', 'Completion Notes', 'Timestamp']
            else:  # electrical or mechanical issues
                headers = ['ID', 'Area', 'Shift', 'Issue Type', 'Description', 'Reported By', 'Resolved', 'Resolution Notes', 'Timestamp']
            
            # Add table headers
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered_alignment
                cell.border = border
                
            ws.row_dimensions[3].height = 20
            
            # Add data rows
            for row_idx, item in enumerate(data, 4):
                ws.cell(row=row_idx, column=1).value = item['id']
                ws.cell(row=row_idx, column=2).value = item['area']
                
                if data_type == 'planned_maintenance':
                    ws.cell(row=row_idx, column=3).value = item['maintenance_type']
                    ws.cell(row=row_idx, column=4).value = item['description']
                    ws.cell(row=row_idx, column=5).value = item['scheduled_date']
                    ws.cell(row=row_idx, column=6).value = item['scheduled_by']
                    ws.cell(row=row_idx, column=7).value = 'Yes' if item['completed'] else 'No'
                    ws.cell(row=row_idx, column=8).value = item['completion_notes']
                    ws.cell(row=row_idx, column=9).value = item['timestamp']
                else:  # electrical or mechanical issues
                    ws.cell(row=row_idx, column=3).value = item['shift']
                    ws.cell(row=row_idx, column=4).value = item['issue_type']
                    ws.cell(row=row_idx, column=5).value = item['description']
                    ws.cell(row=row_idx, column=6).value = item['reported_by']
                    ws.cell(row=row_idx, column=7).value = 'Yes' if item['resolved'] else 'No'
                    ws.cell(row=row_idx, column=8).value = item['resolution_notes']
                    ws.cell(row=row_idx, column=9).value = item['timestamp']
                
                # Apply borders to all cells in the row
                for col_idx in range(1, 10):
                    ws.cell(row=row_idx, column=col_idx).border = border
        
        # Auto-adjust column widths
        for col_idx in range(1, 10):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 15
        
        # Special width adjustments
        ws.column_dimensions['D'].width = 20  # Issue Type/Maintenance Type
        ws.column_dimensions['E'].width = 40  # Description
        ws.column_dimensions['H'].width = 40  # Resolution/Completion Notes
        ws.column_dimensions['I'].width = 20  # Timestamp
        
        # Create a buffer to save the workbook
        excel_buffer = io.BytesIO()
        wb.save(excel_buffer)
        excel_buffer.seek(0)
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            excel_buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=maintenance_{data_type}_{selected_date.strftime("%Y-%m-%d")}.xlsx'
        
        return response
        
    except Exception as e:
        import traceback
        print(f"Excel Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        # If there's an error, return a plain text response with the error
        return HttpResponse(f"Error generating Excel file: {str(e)}", content_type='text/plain')

def export_maintenance_as_pdf(data, selected_date, data_type, title):
    """Export maintenance data as PDF file"""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from django.http import HttpResponse
    import io
    
    # Create a buffer to receive PDF data
    buffer = io.BytesIO()
    
    try:
        # Create the PDF document using the buffer
        doc = SimpleDocTemplate(
            buffer,
            pagesize=landscape(letter),
            rightMargin=30,
            leftMargin=30,
            topMargin=30,
            bottomMargin=30
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = styles['Heading1']
        title_style.alignment = 1  # Center alignment
        normal_style = styles['Normal']
        
        # Create elements to add to the PDF
        elements = []
        
        # Add title
        title_text = f"{title} - {selected_date.strftime('%B %d, %Y')}"
        title_paragraph = Paragraph(title_text, title_style)
        elements.append(title_paragraph)
        elements.append(Spacer(1, 20))
        
        # Skip if no data
        if not data:
            no_data = Paragraph("No data available", normal_style)
            elements.append(no_data)
        else:
            # Define headers based on data type
            if data_type == 'planned_maintenance':
                headers = ['ID', 'Area', 'Type', 'Description', 'Scheduled Date', 'Scheduled By', 'Completed', 'Notes']
            else:  # electrical or mechanical issues
                headers = ['ID', 'Area', 'Shift', 'Issue Type', 'Description', 'Reported By', 'Resolved', 'Notes']
            
            # Prepare table data
            table_data = [headers]
            
            for item in data:
                if data_type == 'planned_maintenance':
                    row = [
                        str(item['id']),
                        item['area'],
                        item['maintenance_type'],
                        item['description'],
                        item['scheduled_date'],
                        item['scheduled_by'],
                        'Yes' if item['completed'] else 'No',
                        item['completion_notes'] or ''
                    ]
                else:  # electrical or mechanical issues
                    row = [
                        str(item['id']),
                        item['area'],
                        item['shift'],
                        item['issue_type'],
                        item['description'],
                        item['reported_by'],
                        'Yes' if item['resolved'] else 'No',
                        item['resolution_notes'] or ''
                    ]
                table_data.append(row)
            
            # Create table
            table = Table(table_data)
            
            # Apply table styles
            table.setStyle(TableStyle([
                # Header styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                
                # Data row styling
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # ID column centered
                
                # Grid styling
                ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.whitesmoke, colors.white]),
                
                # Padding
                ('TOPPADDING', (0, 0), (-1, -1), 3),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
                ('RIGHTPADDING', (0, 0), (-1, -1), 5),
            ]))
            
            # Add table to elements
            elements.append(table)
        
        # Build the PDF document
        doc.build(elements)
        
        # Get the value of the buffer
        pdf_data = buffer.getvalue()
        buffer.close()
        
        # Create HTTP response with PDF file
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename=maintenance_{data_type}_{selected_date.strftime("%Y-%m-%d")}.pdf'
        response.write(pdf_data)
        
        return response
        
    except Exception as e:
        import traceback
        print(f"PDF Generation Error: {str(e)}")
        print(traceback.format_exc())
        
        # If there's an error, return a plain text response with the error
        return HttpResponse(f"Error generating PDF: {str(e)}", content_type='text/plain') 